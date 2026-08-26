## Use this to update

###  Model Training   

from google.cloud import bigquery
import pandas as pd
import numpy as np
from datetime import date, datetime, timedelta
import os
from scripts.gather_historic_data import gather_historic_data 
from xgboost import XGBRegressor
from sklearn.metrics import (mean_absolute_error, mean_squared_error, r2_score)
from openpyxl import load_workbook
from scripts.gather_data_to_forecast import gather_data_to_forecast
from datetime import datetime


from scripts.data_pull_functions import login_google_cloud
from scripts.gather_historic_data import gather_historic_data
from scripts.gather_data_to_forecast import gather_data_to_forecast


client = login_google_cloud(project_name="bepc-prj-energy-prod")

gather_historic_data(client=client, start='2023-08-01', end=datetime.today().strftime("%Y-%m-%d"), save_output=True)

historic_df = pd.read_csv(r"./data/processed-data/historic_data_df.csv")


#gather_historic_data(start='2023-08-01', end=datetime.today().strftime("%Y-%m-%d"), save_output=True)

#historic_df = pd.read_csv(r"C:\VSCode\Workspace\gas-burn-by-site\data\processed-data\historic_data_df.csv")

historic_df["datetime"] = pd.to_datetime(historic_df["datetime"])



#create copy of df then sort
model_df = historic_df.copy()

#sorting by site and datetime to not mix all the sites together
model_df = model_df.sort_values(["site", "datetime"])


#create lag features
model_df["gas_lag_1"] = (model_df.groupby("site")["hourly_gas_burn_MMBtu"].shift(1))

model_df["gas_lag_24"] = (model_df.groupby("site")["hourly_gas_burn_MMBtu"].shift(24))

model_df["gas_lag_168"] = (model_df.groupby("site")["hourly_gas_burn_MMBtu"].shift(168))


#create rolling avgs 
model_df["gas_roll_24"] = (model_df.groupby("site") ["hourly_gas_burn_MMBtu"].transform(lambda x: x.rolling(24).mean()))

model_df["gas_roll_168"] = (model_df.groupby("site")["hourly_gas_burn_MMBtu"].transform(lambda x: x.rolling(168).mean()))


#removing NaNs
model_df = model_df.dropna(subset=[ "gas_lag_1",  "gas_lag_24", "gas_lag_168", "gas_roll_24", "gas_roll_168"])




#grabbing all the features from the datasets
base_features = ["availability_mw", "load_forecast", "net_load_forecast", "wind_forecast", "temperature_forecast", "wind_speed_forecast", "total_offline_forecast", 
           "offline_ng_forecast", "offline_coal_forecast", "hour", "day_of_week", "month"]

features = base_features + ["gas_lag_1", "gas_lag_24", "gas_lag_168", "gas_roll_24", "gas_roll_168"]




model_df["availability_mw"] = (model_df.groupby("site")["availability_mw"].transform(lambda x: x.dropna() ))

models = {}
results = {}

for site in model_df["site"].unique():

    site_df = model_df[ model_df["site"] == site].copy()
   


    #train on the past and test on the future
    train = site_df[site_df["datetime"] < "2026-01-01"].copy()

    test = site_df[site_df["datetime"] >= "2026-01-01"].copy()



    #Build the model
    model = XGBRegressor(n_estimators=500, max_depth=6, learning_rate=0.03, subsample=0.8, colsample_bytree=0.8, random_state=42)


    #Fit the model
    model.fit(train[features], train["hourly_gas_burn_MMBtu"])

    models[site] = model
    
    #generate predictions 
    prediction = model.predict(test[features])


    #evaluate 
    mean_abs_error = mean_absolute_error(test["hourly_gas_burn_MMBtu"], prediction)

    root_mean_squared_error = np.sqrt(mean_squared_error(test["hourly_gas_burn_MMBtu"], prediction))

    r2 = r2_score(test["hourly_gas_burn_MMBtu"], prediction)


    results[site] = {"mae": mean_abs_error, "rmse": root_mean_squared_error, "r2": r2}





##Load Future Forecast Data



forecast_start = datetime.today().strftime("%Y-%m-%d")
forecast_end = (datetime.today() + timedelta(days=7)).strftime("%Y-%m-%d")

gather_data_to_forecast(forecast_start = forecast_start, forecast_end = forecast_end, save_output=True)

forward_df = pd.read_csv("./data/processed-data/data_to_forecast_df.csv")
forward_df["datetime"] = pd.to_datetime(forward_df["datetime"])


# Fill availability the same way as training data

forward_df["availability_mw"] = (forward_df.groupby("site")["availability_mw"].transform(lambda x: x.fillna(x.median())))




##### Recursive Forecast Loop 


all_forecasts = []

for site in models:

    print(f"Forecasting for {site}")

    model = models[site]


    # Future hours for this site
    site_future = (forward_df[forward_df["site"] == site].sort_values("datetime").copy())



    # Last 168 actual hours used to seed lags
    history = (model_df[model_df["site"] == site].sort_values("datetime").tail(168).copy())

    for _, row in site_future.iterrows():

        row = row.copy()


    # Create lag features


        row["gas_lag_1"] = (history["hourly_gas_burn_MMBtu"].iloc[-1])

        row["gas_lag_24"] = (history["hourly_gas_burn_MMBtu"].iloc[-24])

        row["gas_lag_168"] = (history["hourly_gas_burn_MMBtu"].iloc[-168])

        row["gas_roll_24"] = (history["hourly_gas_burn_MMBtu"].tail(24).mean())

        row["gas_roll_168"] = (history["hourly_gas_burn_MMBtu"].tail(168).mean())





    # Make prediction
        X_pred = pd.DataFrame([row])[features]

        if row["availability_mw"] <= 1:

            prediction = 0
            
        else:

            prediction = model.predict(X_pred)[0]
            
        prediction = max(0, prediction)
        



# Save forecast

        all_forecasts.append({"datetime": row["datetime"], "site": site, "predicted_gas_burn_MMBtu": prediction})

        forecast_df = pd.DataFrame(all_forecasts)




# Add prediction back into history for future lag values

    history = pd.concat([history, pd.DataFrame([{"datetime": row["datetime"], "hourly_gas_burn_MMBtu": prediction}])], ignore_index=True)

    # Keep most recent 168 hours (7 days)
    history = history.tail(168)


#Final Forecast DataFrame

print("forecast records:", len(all_forecasts))
forecast_df = pd.DataFrame(all_forecasts)



# Create NBPL layout
format = (forecast_df.pivot(index="datetime", columns="site", values="predicted_gas_burn_MMBtu").reset_index())

# Only first 24 hours belong on the burn sheet
format = format.head(24)

print(format.head())
#change this to a download the user has

template_file = r"G:\Trading\Forecasts\Daily Gas Burn Forecast by Site\Forecasts\Forecast template.xlsx"

workbook = load_workbook(template_file, keep_vba=True)

worksheet = workbook["Daily Burn Sheet"]

        
gas_day = format["datetime"].min().date()

worksheet["D11"] = gas_day
worksheet["F11"] = gas_day
 
        
for i, (_, row) in enumerate(format.iterrows()):

    excel_row = 30 + i

    worksheet.cell(row=excel_row, column=6, value=row["CGS"])   # F
    worksheet.cell(row=excel_row, column=8, value=row["DCS"])   # H
    worksheet.cell(row=excel_row, column=10, value=row["GGS"])  # J
    worksheet.cell(row=excel_row, column=12, value=row["LCS"])  # L
    worksheet.cell(row=excel_row, column=14, value=row["PGS"])  # N



today_str = datetime.today().strftime("%m.%d.%Y")

output_file = (rf"G:\Trading\Forecasts\Daily Gas Burn Forecast by Site\Forecasts\Hourly Gas Burn Forecast for {today_str}.xlsm")

workbook.save(output_file)

print(output_file)