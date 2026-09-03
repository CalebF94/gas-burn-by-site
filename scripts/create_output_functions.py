"""
Module for creating formatted Excel output files from gas burn predictions.

Functions:
    - earliest_full_gas_day_with_forecast: Identify the earliest gas day with complete forecast coverage
    - create_nbpl_file: Generate NBPL (Northern Border Pipeline) Excel forecast file
    - create_next_day_gas_burn_file: Generate next-day gas burn summary and hourly detail report
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from scripts.constants import NBPL_TEMPLATE, NEXT_DAY_GAS_BURN_TEMPLATE


def earliest_full_gas_day_with_forecast(predictions: pd.DataFrame) -> pd.DataFrame:
    """
    Identify the earliest gas day with complete forecast coverage across all sites.
    
    Determines which gas days have forecasts for all sites by counting the number of unique
    sites for each gas day and finding days where the site count matches the maximum count
    (i.e., all sites are represented).

    Parameters:
        predictions: DataFrame containing forecast records with 'gasday' and 'site' columns.
                     Must include at least one row per site per gas day for complete coverage.

    Returns:
        pd.DatetimeIndex: DatetimeIndex of gas days with complete site coverage (all sites
                          present for that day). Results are converted to datetime format.
    """
    daily_record_counts = predictions.groupby('gasday').count()['site'].reset_index()
    full_gas_days = daily_record_counts[daily_record_counts['site']==max(daily_record_counts['site'])]['gasday']
    full_gas_days = pd.to_datetime(full_gas_days)

    return full_gas_days

def create_nbpl_file(predictions: pd.DataFrame, template_file: str = None, file_date: str = None, save_file: str = None) -> None:
    """
    Generate NBPL (Northern Border Pipeline) Excel forecast file with hourly gas burn predictions.
    
    Populates an Excel template with hourly gas burn forecasts for each site organized by hour-ending.
    The output is an Excel file with VBA macros intact containing the forecast data formatted for
    NBPL submission.

    Parameters:
        predictions: DataFrame containing site-level hourly gas burn predictions with columns:
                     'site', 'gasday', 'hourly_gas_burn_MMBtu'
        template_file: Path to Excel template file. If None, uses default NBPL template.
                       Default is None.
        file_date: Date string to use for the forecast gas day. If None, uses the earliest gas day
                   with complete predictions across all sites. Default is None.
        save_file: Path where to save the output Excel file. If None, saves to default location
                   with today's date in filename. Default is None.

    Returns:
        None

    Notes:
        - The output excel file is saved to the G drive by default unless a different path is specified in save_file.
        - Creates/updates cells D11 and F11 with the forecast date
        - Populates hourly values (rows 30-53) for each site in columns mapped to CGS, DCS, GGS, LCS, PGS
        - Row 55 contains SUM formulas for daily totals
        - Preserves VBA macros from the template file
    """
    # cleaning up file specifications
    today_str = datetime.today().strftime("%m.%d.%Y")
    template_str = template_file if template_file else Path(NBPL_TEMPLATE)
    save_file_str = save_file if save_file else Path(rf"G:\Trading\Forecasts\Daily Gas Burn Forecast by Site\Forecasts\Hourly Gas Burn Forecast for {today_str}.xlsm")

    template_path = Path(template_str)
    save_path = Path(save_file_str)


    # dynamically getting date based on first day with full gas day predictions
    if not file_date:
        full_gas_days = earliest_full_gas_day_with_forecast(predictions=predictions)
        file_date = min(full_gas_days).date()

    else:
        file_date = pd.to_datetime(file_date)


    try:
        workbook = load_workbook(template_path, keep_vba=True)
        worksheet = workbook["Daily Burn Sheet"]

        # updating start and end date in the file
        worksheet["D11"] = file_date
        worksheet["F11"] = file_date

        site_to_columns = {'CGS': 6, 'DCS': 8, 'GGS': 10, 'LCS': 12, 'PGS': 14}
            
        for site, col in site_to_columns.items():
            worksheet.cell(row=55, column=col, value=f'=SUM({chr(64+col)}30:{chr(64+col)}53)')
            site_predictions = predictions.loc[(predictions['site']==site) & (predictions['gasday']==file_date), 'hourly_gas_burn_MMBtu']
            for row in range(0, 24):
                
                excel_row = 30 + row
                worksheet.cell(row=excel_row, column=col, value=site_predictions.values[row])

        workbook.save(save_path)
        print(f'A file with the NBPL hourly forecasts is saved to {save_path}')

    finally:
        if hasattr(workbook, 'vba_archive') and workbook.vba_archive:
            workbook.vba_archive.close()
            
        workbook.close()

    


def create_next_day_gas_burn_file(predictions: pd.DataFrame, template_file: str = None, save_file: str = None) -> None:
    """
    Generate next-day gas burn summary and hourly detail report Excel file.
    
    Creates an Excel workbook with two worksheets: a summary sheet with daily gas burn by site
    and an hourly detail sheet with hourly gas burn values pivoted by hour-ending for each site.
    Automatically calculates column widths for readability.

    Parameters:
        predictions: DataFrame containing hourly gas burn predictions with columns:
                     'site', 'date', 'gasday', 'gasday_of_week', 'HE', 'effective_day_of_week',
                     'hourly_gas_burn_MMBtu'
        template_file: Path to Excel template file. If None, uses default template.
                       Default is None.
        save_file: Path where to save the output Excel file. If None, saves to default location
                   with today's date in filename. Default is None.

    Returns:
        None

    Notes:
        - Summary sheet: Shows daily totals by site and day of week (rows 3-10, columns 3-10)
        - Hourly sheet: Shows hourly forecasts by site with hour columns (HE1-HE24) and date rows
        - Only includes forecasted days that are in the future (after current time)
        - Automatically formats date cells as mm/dd/yyyy and numeric cells with one decimal place
        - Auto-adjusts column widths for all columns in hourly sheet
    """

    # cleaning up file specifications
    today_str = datetime.today().strftime("%m.%d.%Y")
    template_str = template_file if template_file else Path(NEXT_DAY_GAS_BURN_TEMPLATE)
    save_file_str = save_file if save_file else Path(f'G:/Trading/Forecasts/Daily Gas Burn Forecast by Site/Next Day Gas Burn Files/Next Day Gas Burn {today_str}.xlsx')

    template_path = Path(template_str)
    save_path = Path(save_file_str)


    #load file
    try:
        next_day_gas_burn_wb = load_workbook(template_path, data_only=True)
        summary_ws = next_day_gas_burn_wb['Summary']
        hourly_gas_use_ws = next_day_gas_burn_wb['Hourly Gas Use']

        #identifies which days from predictions df will be on report
        forecasted_days = earliest_full_gas_day_with_forecast(predictions)
        forecasted_days = forecasted_days[forecasted_days >  pd.Timestamp.now()]


        #update summary tab
        formatted_predictions = (
        predictions
        .drop(columns=['datetime', 'date','HE'])
        .groupby(by=['gasday', 'gasday_of_week', 'site'])
        .sum()
        .reset_index()
        .pivot(columns='site', index=['gasday', 'gasday_of_week'], values='hourly_gas_burn_MMBtu')
        .round(0)
        .reset_index()
        )

        formatted_predictions = formatted_predictions.loc[formatted_predictions['gasday'].isin(forecasted_days)]
        formatted_predictions = formatted_predictions.reindex(columns=['gasday', 'gasday_of_week', 'DCS', 'GGS', 'CGS', 'PGS', 'LCS'])

        # range should contain 7 days for weekly forecast. Starts on row 3 in template. Limiting to shape[0] to avoid errors if a partial week is needed
        for df_row, excel_row in enumerate(range(3, 3+formatted_predictions.shape[0])): 
            for df_col, excel_col in enumerate(range(3, 10)):
                summary_ws.cell(row=excel_row, column=excel_col, value=formatted_predictions.iloc[df_row, df_col])
                #print(formatted_predictions.iloc[df_row, df_col])


        #update hourly gas use tab
        hourly_predictions_formatted = (
        predictions
        .drop(columns=['gasday', 'datetime', 'gasday_of_week'])
        .pivot(index=['site', 'date', 'effective_day_of_week'], columns = 'HE', values='hourly_gas_burn_MMBtu')
        .reset_index()
        .round(2)
        )

        site_start_row = {'GGS': 3, 'CGS': 13, 'DCS': 23, 'PGS': 33, 'LCS': 43}

        for site in site_start_row:
            site_predictions = hourly_predictions_formatted[hourly_predictions_formatted['site']==site]
            start_row = site_start_row[site]
            for df_row, excel_row in enumerate(range(start_row, start_row + site_predictions.shape[0])):
                
                for df_col, excel_col in enumerate(range(4, 30)): # range includes 24 hours (HE1 to HE24) plus two date columns

                    hourly_gas_use_ws.cell(row=excel_row, column=excel_col, value=site_predictions.iloc[df_row, df_col+1])#+1 is b/c Excel is 1-indexed and df is 0-indexed.
                    if excel_col == 4: 
                        hourly_gas_use_ws.cell(row=excel_row, column=excel_col).number_format = "mm/dd/yyyy"
                    else:  
                        hourly_gas_use_ws.cell(row=excel_row, column=excel_col).number_format = "#,##0.0"


        #widening columns
        for col in hourly_gas_use_ws.columns:
            max_len = 0
            col_letter = col[0].column_letter  # Get letter like 'A', 'B', etc.
            for cell in col:
                if cell.value is not None:
                    # Check length of the cell string
                    max_len = max(max_len, len(str(cell.value)))
                    
            # Add a little padding (e.g., +3) so it isn't too tight
            hourly_gas_use_ws.column_dimensions[col_letter].width = max(max_len + .1, 10)

        # saving file
        next_day_gas_burn_wb.save(save_path)
        print(f'A file with the next day gas burns is saved to {save_path}')

    finally:
        next_day_gas_burn_wb.close()
